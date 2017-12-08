import openpyxl
import tkinter.filedialog as Tkinter
import os.path
import codecs as loscodecs

def ConseguirStubs():
    archivoTXT = Tkinter.askopenfilename(initialdir = "C:/Users/IBM_ADMIN/Desktop/archivosLaYessenia/",
                                         title = "Archivos TXT")
    archivo = loscodecs.open(archivoTXT, "r")
    dicProgramasStubs = {}

    key = ""
    value = ""

    for linea in archivo:
        lineaStrip = linea.strip()

        if lineaStrip[:6] == "Stub: ":
            value = lineaStrip[6:]

        elif lineaStrip[:10] == "Programa: ":
            key = (lineaStrip[10:])

        dicProgramasStubs[key] = value
    print(dicProgramasStubs)
    archivo.close()

    InsertarStubs(dicProgramasStubs)

def InsertarStubs(programas):
    # Seleccionar el Excel a usar
    archivo = Tkinter.askopenfilename(initialdir = "C:/Users/IBM_ADMIN/PycharmProjects/laYesseniaDemo/", title = "Archivo Excel")

    # Abrirlo usando la libreria openpyxl, en este caso 'keep_vba' es para mantener los codigos visual basic que se han utilizado
    workbook = openpyxl.load_workbook(os.path.basename(str(archivo)), keep_vba = True)
    sheet = workbook["BD_Tesoreria"]

    for row in sheet.iter_cols(min_row = 1, min_col = 1, max_row = 4125, max_col = 11):
        for cell in row:
            programa = str(cell.value)
            if programa[:8] in programas.keys():
                programasKeys = programas[programa[:8]]
                stubs = "G" + str(cell.row)
                print(stubs, programasKeys)
                sheet[stubs] = programasKeys

    workbook.save("lonotarteamiymetiraspasando.xlsm")

def ConseguirEntradaSalida():
    archivoTXT = Tkinter.askopenfilename(initialdir = "C:/Users/IBM_ADMIN/Desktop/archivosLaYessenia/",
                                         title = "Archivos TXT")
    archivo = loscodecs.open(archivoTXT, "r")
    dicEntradaSalida = {}

    key = ""
    value = []

    for linea in archivo:
        lineaStrip = linea.strip()

        if lineaStrip == "":
            if key != "":
                InsertarEntradaSalida(dicEntradaSalida)

            dicEntradaSalida.clear()
            key = ""
            value.clear()

        elif lineaStrip[:13] == "StubEntrada: ":
            value.append(lineaStrip[13:])

        elif lineaStrip[:12] == "StubSalida: ":
            value.append(lineaStrip[12:])

        elif lineaStrip[:6] == "Stub: ":
            key = (lineaStrip[6:])

        dicEntradaSalida[key] = value

    archivo.close()

def InsertarEntradaSalida(dicEntradaSalida):
    workbook = openpyxl.load_workbook("Modulo_de_Aplicacion_Tesoreria-Operativa_Terminado.xlsm")
    sheet = workbook["BD_Tesoreria"]
    listAyuda = []

    for row in sheet.iter_cols(min_row = 1, min_col = 1, max_row = 4125, max_col = 11):
        for cell in row:
            entradaSalida = str(cell.value)
            if entradaSalida in dicEntradaSalida.keys():
                entrada = "H" + str(cell.row)
                salida = "I" + str(cell.row)

                for valores in dicEntradaSalida[entradaSalida]:
                    celdaEntrada = sheet[entrada]
                    #sheet[salida] = valores


                #for valores in dicEntradaSalida[cell.value]:
                    #print(valores)
                '''
                    if valores[:5] == 'Ktoli':
                        Ktoli = valores[:5]
                        entrada = "H" + str(cell.row)
                        print(entrada, dicEntradaSalida)
                            #if valores2.startswith(Ktoli):
                                #print(entrada, ": ", dicEntradaSalida[Ktoli])
                       #sheet[entrada] = entradaSalidaValores
                    elif valores[:5] == 'ktolv':
                        ktolv = valores[:5]
                        salida = "I" + str(cell.row)
                        print(entradaSalidaValores[ktolv])
                        #sheet[salida] = entradaSalidaValores
                    '''

    workbook.save("lonotarteamiymetiraspasando2.xlsx")

ConseguirEntradaSalida()
#ConseguirStubs()